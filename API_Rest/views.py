from django.shortcuts import render, redirect
from django.urls import reverse
from .models import Producto, Servicio, Cuenta_Empleado,Cliente
from openpyxl import Workbook
from django.http import HttpResponse
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.hashers import check_password
from .models import Cuenta_Empleado
from django.contrib.auth.hashers import make_password
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from django.core.serializers import serialize


def crud_productos(request):
    productos = Producto.objects.all()
    return render(request, "core/crud_productos.html", {"productos": productos})

def crud_empleados(request):
    cuenta_empleado = Cuenta_Empleado.objects.all()
    return render(request, "core/crud_cuentas.html", {"cuenta_empleado": cuenta_empleado})


def index(request):
    return render(request, "core/index.html")

def servicios(request):
    
    servicios = Servicio.objects.all()

    return render(request, 'core/servicios.html', {"servicios": servicios})


def perfil(request):
    return render(request, "core/perfil.html")

def productos(request):
    productos = Producto.objects.all()
    return render(request, "core/productos.html", {"productos": productos})

def deshabilitar_producto(request, id_producto):
    # Buscar el producto por su ID
    producto = Producto.objects.get(pk=id_producto)

    # Cambiar el estado de habilitado a False
    producto.habilitado = False
    producto.save()

    # Redirigir a la página de productos
    return redirect('productos')

def habilitar_producto(request, Id_Producto):
    producto = Producto.objects.get(Id_Producto=Id_Producto)
    producto.habilitado = True
    producto.save()
    return redirect('crud_productos')

def Validar_Pago(request):     
    return render(request, "core/Validar_Pago.html")  

def detalle_venta(request):     
    return render(request, "core/detalle_venta.html")

def pedidos(request):
    return render(request, "core/pedidos.html")

def login(request):
    
    return render(request, "core/login.html")

def MenuEmpleados(request):
    return render(request, "core/MenuEmpleados.html")

def MenuAdmin(request):
    return render(request, "core/MenuAdmin.html")





def log_out(request):
    return render(request, "core/login.html")


def registro(request):
    clientes = Cliente.objects.all()
    return render(request, "core/registro.html", {"clientes": clientes})




def Generar_informes(request):
    return render(request, "core/Generar_Informes.html")

def crud_Servicios(request):
    servicios = Servicio.objects.all()
    return render(request, "core/crud_Servicios.html", {"servicios": servicios})

def crud_cuentas(request):
    cuenta_empleado = Cuenta_Empleado.objects.all()
    cliente = Cliente.objects.all()
    return render(request, "core/crud_cuentas.html", {"cuenta_empleado": cuenta_empleado})


def EliminarCuenta(request, Id_Empleado):
    cuenta = Cuenta_Empleado.objects.get(Id_Empleado=Id_Empleado)
    cuenta.delete()

    return redirect(reverse("crud_cuentas"))

def registrarEmpleado(request):
    if request.method == 'POST':
        Id_Empleado = request.POST["txtId_Empleado"]
        Primer_Nombre = request.POST["txtPrimer_nombre_Empleado"]
        Segundo_Nombre = request.POST["txtSegundo_nombre_Empleado"]
        Primer_Apellido = request.POST["txtPrimer_Apellido"]
        Segundo_Apellido = request.POST["txtSegundo_Apellido"]
        Direccion = request.POST["txtDireccion"]
        Edad = request.POST["txtEdad"]
        Cargo = request.POST["txtCargo"]

        # Generar usuario y contraseña según las especificaciones
        usuario = Primer_Nombre[:2].lower() + '.' + Segundo_Nombre[-2:].lower() + '_' +Edad
        contrasena = usuario  # Establecer la contraseña inicial como el usuario


        cuenta_empleado = Cuenta_Empleado.objects.create(
            Id_Empleado=Id_Empleado,
            Primer_Nombre=Primer_Nombre,
            Segundo_Nombre=Segundo_Nombre,
            Primer_Apellido=Primer_Apellido,
            Segundo_Apellido=Segundo_Apellido,
            Direccion=Direccion,
            Edad=Edad,
            Cargo=Cargo,
            Usuario=usuario,
            Contrasena=contrasena # Guardar la contraseña hasheada
        )
        cuenta_empleado.save()
        messages.success(request, 'Cuenta de empleado registrada exitosamente.')
        return redirect(reverse('crud_cuentas'))


def registrarCliente(request):
    if request.method == 'POST':
        nombre_usuario = request.POST.get("txtNombre_Usuario")
        contrasena = request.POST.get("txtContrasena")
        correo_electronico = request.POST.get("txtCorreoC")
        edad = request.POST.get("txtAge")
        comuna = request.POST.get("ComunaC")
        numero_telefonico = request.POST.get("txtNumberTelephone")
        genero = request.POST.get("radioMasculino") or request.POST.get("radioFemenino") or request.POST.get("radioOtros")

        cliente = Cliente.objects.create(
            Nombre_Usuario=nombre_usuario,
            Contrasena=contrasena,
            Correo_Electronico=correo_electronico,
            Edad=edad,
            Comuna=comuna,
            Numero_Telefonico=numero_telefonico,
            Genero=genero
        )

        cliente.save()
        messages.success(request, 'Cuenta de cliente registrada exitosamente.')
        return redirect('registro')

    return render(request, 'core/registro.html')


def eliminarProducto(request, Id_Producto):
    producto = Producto.objects.get(Id_Producto=Id_Producto)
    producto.delete()

    return redirect(reverse("crud_productos"))

def registrarProductos(request):
    id_producto = request.POST["txtId_Producto"]
    nombre_producto = request.POST["txtNombre_Producto"]
    stock_producto = request.POST["txtstock_Producto"]
    descripcion_producto = request.POST["txtdescripcion_Producto"]
    precio = request.POST["txtPrecio_Producto"]
    categoria = request.POST["txtCategoria_Producto"]
    Marca = request.POST["txtMarca_Producto"]
    Proveedor = request.POST["txtProveedor_Producto"]

    producto = Producto()
    producto.Id_Producto = id_producto
    producto.Nombre_Producto = nombre_producto
    producto.stock_Producto = stock_producto
    producto.descripcion_Producto = descripcion_producto
    producto.Precio_Producto = precio
    producto.Categoria_Producto = categoria
    producto.Marca_Producto = Marca
    producto.Proveedor_Producto = Proveedor

    producto.save()

    return redirect(reverse("crud_productos"))

def EditarProductos(request):
    id_producto = request.POST["txtId_Producto"]
    nombre_producto = request.POST["txtNombre_Producto"]
    stock_producto = request.POST["txtstock_Producto"]
    descripcion_producto = request.POST["txtdescripcion_Producto"]
    precio = request.POST["txtPrecio_Producto"]
    categoria = request.POST["txtCategoria_Producto"]
    Marca = request.POST["txtMarca_Producto"]
    Proveedor = request.POST["txtProveedor_Producto"]

    producto = Producto.objects.get(Id_Producto=id_producto)
    producto = Producto()
    producto.Id_Producto = id_producto
    producto.Nombre_Producto = nombre_producto
    producto.stock_Producto = stock_producto
    producto.descripcion_Producto = descripcion_producto
    producto.Precio_Producto = precio
    producto.Categoria_Producto = categoria
    producto.Marca_Producto = Marca
    producto.Proveedor_Producto = Proveedor
    producto.save()

    return redirect(reverse("crud_productos"))




def registrarServicios(request):
    id_servicio = request.POST["txtId_Servicio"]
    nombre_servicio = request.POST["txtNombre_Servicio"]
    tipo_servicio = request.POST["txtTipo_Servicio"]
    precio_servicio = request.POST["txtPrecio_Servicio"]
    personal_cargo = request.POST["txtPersonal_cargo"]

    servicio = Servicio()
    servicio.Id_Servicio = id_servicio
    servicio.Nombre_Servicio = nombre_servicio
    servicio.Tipo_Servicio = tipo_servicio
    servicio.Precio_Servicio = precio_servicio
    servicio.Personal_cargo = personal_cargo

    servicio.save()

    return redirect(reverse("crud_Servicios"))


def edicionServicio(request, Id_Servicio):
    servicio = Servicio.objects.get(Id_Servicio=Id_Servicio)
    return render(request, "core/edicionServicio.html", {"servicio": servicio})


def EliminarServicio(request, Id_Servicio):
    servicio = Servicio.objects.get(Id_Servicio=Id_Servicio)
    servicio.delete()

    return redirect(reverse("crud_Servicios"))







def edicionProducto(request, Id_Producto):
    producto = Producto.objects.get(Id_Producto=Id_Producto)
    return render(request, "core/edicionProducto.html", {"producto": producto})

def EditarServicios(request):
    id_servicio = request.POST["txtId_Servicio"]
    nombre_servicio = request.POST["txtNombre_Servicio"]
    tipo_servicio = request.POST["txtTipo_Servicio"]
    precio_servicio = request.POST["txtPrecio_Servicio"]
    personal_cargo = request.POST["txtPersonal_cargo"]

    servicio = Servicio.objects.get(Id_Servicio=id_servicio)
    servicio = Servicio()
    servicio.Id_Servicio = id_servicio
    servicio.Nombre_Servicio = nombre_servicio
    servicio.Tipo_Servicio = tipo_servicio
    servicio.Precio_Servicio = precio_servicio
    servicio.Personal_cargo = personal_cargo
    servicio.save()

    return redirect(reverse("crud_Servicios"))



def generar_informe(request):
    # Lógica para obtener datos de productos, servicios y cuentas
    productos = Producto.objects.all()
    servicios = Servicio.objects.all()
    cuentas = Cuenta_Empleado.objects.all()

    # Crear un libro de Excel
    wb = Workbook()

    # Crear hojas para productos, servicios y cuentas
    ws_productos = wb.create_sheet(title="Productos")
    ws_servicios = wb.create_sheet(title="Servicios")
    ws_cuentas = wb.create_sheet(title="Cuentas")

    # Escribir datos en las hojas
    for producto in productos:
        ws_productos.append([producto.Id_Producto, producto.Nombre_Producto, producto.stock_Producto, producto.descripcion_Producto, producto.Precio_Producto, producto.Categoria_Producto, producto.Marca_Producto, producto.Proveedor_Producto])

    for servicio in servicios:
        ws_servicios.append([servicio.Id_Servicio, servicio.Nombre_Servicio, servicio.Tipo_Servicio, servicio.Precio_Servicio, servicio.Personal_cargo])

    for cuenta in cuentas:
        ws_cuentas.append([cuenta.Id_Empleado, cuenta.Primer_Nombre, cuenta.Segundo_Nombre, cuenta.Primer_Apellido, cuenta.Segundo_Apellido, cuenta.Direccion, cuenta.Edad, cuenta.Cargo])

    # Crear una respuesta de Django para el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=informe.xlsx'

    # Guardar el libro de Excel en la respuesta
    wb.save(response)
    
    
    return response

def loginEmpleado(request):
    cuentas_empleados = serialize('json', Cuenta_Empleado.objects.all())
    return render(request, 'core/LoginEmpleados.html', {'cuentas_empleados': cuentas_empleados})

















def LoginEmpleados(request):
    return render(request, "core/LoginEmpleados.html")